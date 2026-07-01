import streamlit as st
import geopandas as gpd
import pandas as pd
import folium
from folium.plugins import MarkerCluster
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile, os, io, tempfile
from datetime import datetime

# ── Configuração da página ─────────────────────────────────────
st.set_page_config(
    page_title="Visualizador de Redes — Águas do Rio",
    page_icon="🗺️",
    layout="wide",
)

st.markdown("""
<style>
    .titulo { color: #1F4E79; font-size: 2rem; font-weight: bold; }
    .subtitulo { color: #2E75B6; font-size: 1rem; margin-bottom: 1rem; }
    div[data-testid="metric-container"] {
        background: white;
        border: 1px solid #E0E8F0;
        border-radius: 8px;
        padding: 0.8rem 1rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="titulo">🗺️ Visualizador de Camadas Vetoriais</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitulo">Águas do Rio | Planejamento e Controle — Visualize atributos e mapa de qualquer camada vetorial</div>', unsafe_allow_html=True)
st.divider()


# ── Funções ────────────────────────────────────────────────────

def ler_arquivo(arquivo_bytes, nome_arquivo):
    ext = os.path.splitext(nome_arquivo)[1].lower()
    tmp_dir = tempfile.mkdtemp()

    if ext == '.zip':
        with zipfile.ZipFile(io.BytesIO(arquivo_bytes)) as z:
            z.extractall(tmp_dir)
        caminho = None
        for root, _, fs in os.walk(tmp_dir):
            for f in fs:
                if f.endswith('.gpkg'):
                    caminho = os.path.join(root, f); break
            if caminho: break
        if not caminho:
            for root, _, fs in os.walk(tmp_dir):
                for f in fs:
                    if f.endswith('.shp'):
                        caminho = os.path.join(root, f); break
                if caminho: break
        if not caminho:
            raise ValueError("Nenhum .gpkg ou .shp encontrado no ZIP.")
    elif ext in ('.gpkg', '.shp'):
        caminho = os.path.join(tmp_dir, nome_arquivo)
        with open(caminho, 'wb') as f:
            f.write(arquivo_bytes)
    else:
        raise ValueError(f"Formato não suportado: {ext}")

    ext_real = os.path.splitext(caminho)[1].lower()
    if ext_real == '.gpkg':
        from pyogrio import list_layers
        camadas = [l[0] for l in list_layers(caminho)]
    else:
        camadas = [os.path.basename(caminho)]

    return caminho, camadas, ext_real


def carregar_camada(caminho, ext, camada):
    if ext == '.gpkg':
        return gpd.read_file(caminho, layer=camada, engine='pyogrio')
    return gpd.read_file(caminho, engine='pyogrio')


PALETA = [
    '#E74C3C','#27AE60','#2980B9','#F39C12','#8E44AD',
    '#16A085','#D35400','#2C3E50','#1ABC9C','#C0392B',
    '#2ECC71','#3498DB','#E67E22','#9B59B6','#1F4E79',
]

def gerar_mapa(gdf, campo_cor):
    gdf_wgs = gdf.to_crs(epsg=4326) if gdf.crs and gdf.crs.to_epsg() != 4326 else gdf.copy()
    centro = [gdf_wgs.geometry.centroid.y.mean(), gdf_wgs.geometry.centroid.x.mean()]
    m = folium.Map(location=centro, zoom_start=13, tiles='CartoDB positron')

    # Montar paleta de cores por valor do campo
    cor_map = {}
    if campo_cor and campo_cor in gdf_wgs.columns:
        valores = gdf_wgs[campo_cor].dropna().unique()
        for i, v in enumerate(sorted([str(x) for x in valores])):
            cor_map[v] = PALETA[i % len(PALETA)]

    cols_tooltip = [c for c in gdf_wgs.columns if c != 'geometry'][:8]

    geom_tipo = gdf_wgs.geometry.geom_type.iloc[0] if len(gdf_wgs) > 0 else 'Unknown'

    if 'Point' in geom_tipo:
        cluster = MarkerCluster(name='Feições').add_to(m)
        for _, row in gdf_wgs.iterrows():
            if row.geometry is None or row.geometry.is_empty: continue
            val = str(row.get(campo_cor, '')) if campo_cor else ''
            cor = cor_map.get(val, '#2980B9') if campo_cor else '#2980B9'
            popup_html = '<br>'.join([f'<b>{c}:</b> {row.get(c,"")}' for c in cols_tooltip])
            folium.CircleMarker(
                location=[row.geometry.y, row.geometry.x],
                radius=6, color=cor, fill=True, fill_color=cor, fill_opacity=0.8,
                popup=folium.Popup(popup_html, max_width=300),
            ).add_to(cluster)
    else:
        # Linhas e polígonos — agrupar por valor do campo
        if campo_cor and cor_map:
            grupos = {}
            for _, row in gdf_wgs.iterrows():
                if row.geometry is None or row.geometry.is_empty: continue
                val = str(row.get(campo_cor, 'Sem valor'))
                cor = cor_map.get(val, '#AAAAAA')
                popup_html = '<br>'.join([f'<b>{c}:</b> {row.get(c,"")}' for c in cols_tooltip])
                feat = folium.GeoJson(
                    row.geometry.__geo_interface__,
                    style_function=lambda f, c=cor: {'color': c, 'weight': 2.5, 'opacity': 0.85,
                                                      'fillColor': c, 'fillOpacity': 0.4},
                    popup=folium.Popup(popup_html, max_width=300),
                )
                if val not in grupos:
                    grupos[val] = folium.FeatureGroup(name=f"{campo_cor}: {val}")
                feat.add_to(grupos[val])
            for g in grupos.values():
                g.add_to(m)
        else:
            folium.GeoJson(
                gdf_wgs.__geo_interface__,
                name='Feições',
                style_function=lambda f: {'color': '#1F4E79', 'weight': 2, 'opacity': 0.8,
                                           'fillColor': '#2E75B6', 'fillOpacity': 0.3},
                tooltip=folium.GeoJsonTooltip(fields=cols_tooltip, labels=True, sticky=False),
            ).add_to(m)

    # Legenda de cores
    if campo_cor and cor_map:
        itens = ''.join([
            f'<div><span style="display:inline-block;width:14px;height:14px;'
            f'background:{c};border-radius:2px;margin-right:6px;"></span>{v}</div>'
            for v, c in list(cor_map.items())[:15]
        ])
        if len(cor_map) > 15:
            itens += f'<div style="color:#888;font-size:11px;">... e mais {len(cor_map)-15} valores</div>'
        legenda = f"""
        <div style="position:fixed;bottom:30px;left:30px;z-index:1000;
             background:white;padding:12px 16px;border-radius:8px;
             border:2px solid #ccc;font-family:Arial;font-size:12px;max-height:300px;overflow-y:auto;">
          <b>🎨 {campo_cor}</b><br><br>{itens}
        </div>"""
        m.get_root().html.add_child(folium.Element(legenda))

    folium.LayerControl().add_to(m)
    return m


def gerar_excel(gdf, nome_arquivo, camada):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Atributos'

    cols = [c for c in gdf.columns if c != 'geometry']
    AZUL = '1F4E79'
    lado = Side(style='thin', color='BFBFBF')
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)

    # Header
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c.upper().replace('_', ' '))
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=AZUL)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borda
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(c) + 4)
    ws.row_dimensions[1].height = 22

    # Dados
    for i, (_, row) in enumerate(gdf.iterrows(), 2):
        fill_cor = 'EBF3FB' if i % 2 == 0 else 'FFFFFF'
        fill = PatternFill('solid', start_color=fill_cor)
        for j, c in enumerate(cols, 1):
            val = row.get(c, '')
            if hasattr(val, 'item'): val = val.item()
            cell = ws.cell(i, j, str(val) if val is not None else '')
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.border = borda
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Aba de metadados
    ws2 = wb.create_sheet('Metadados')
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    meta = [
        ['METADADOS DA CAMADA'],
        ['Arquivo:', nome_arquivo],
        ['Camada:', camada],
        ['CRS:', str(gdf.crs)],
        ['Total de feições:', len(gdf)],
        ['Total de colunas:', len(cols)],
        ['Colunas:', ', '.join(cols)],
        ['Exportado em:', agora],
    ]
    for i, linha in enumerate(meta, 1):
        for j, val in enumerate(linha, 1):
            cell = ws2.cell(i, j, val)
            cell.font = Font(name='Arial', bold=(j == 1 or i == 1), size=11 if i == 1 else 10,
                             color=AZUL if i == 1 else '000000')
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════

# ── Upload ──
st.subheader("📂 Upload do arquivo")
st.caption("Formatos aceitos: **.gpkg** · **.zip** (shapefile ou gpkg dentro)")

arq = st.file_uploader("Selecione o arquivo", type=['gpkg', 'zip', 'shp'],
                        label_visibility='collapsed')

if arq:
    try:
        with st.spinner("Lendo arquivo..."):
            caminho, camadas, ext = ler_arquivo(arq.read(), arq.name)

        # Seleção de camada
        if len(camadas) > 1:
            camada = st.selectbox("📋 Camada do GeoPackage:", camadas)
        else:
            camada = camadas[0]

        with st.spinner("Carregando camada..."):
            gdf = carregar_camada(caminho, ext, camada)

        cols_atrib = [c for c in gdf.columns if c != 'geometry']
        geom_tipo  = gdf.geometry.geom_type.value_counts().index[0] if len(gdf) > 0 else 'N/A'
        n_feicoes  = len(gdf)
        n_cols     = len(cols_atrib)

        st.success(f"✅ **{arq.name}** — camada: `{camada}` carregada com sucesso.")

        # ── Métricas ──
        st.subheader("📊 Resumo da camada")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Feições",    f"{n_feicoes:,}".replace(',', '.'))
        c2.metric("Colunas",    n_cols)
        c3.metric("Geometria",  geom_tipo)
        c4.metric("CRS",        str(gdf.crs.to_epsg()) if gdf.crs else "N/A")

        st.divider()

        # ── Tabs: Tabela | Mapa | Download ──
        tab1, tab2, tab3 = st.tabs(["📋 Tabela de Atributos", "🗺️ Mapa", "📥 Downloads"])

        # ── Tab Tabela ──
        with tab1:
            col_busca, col_info = st.columns([3, 1])
            with col_busca:
                busca = st.text_input("🔍 Filtrar tabela (busca em todos os campos):", placeholder="Digite para filtrar...")
            with col_info:
                st.markdown(f"<br><span style='color:#666;font-size:13px'>{n_feicoes} feições · {n_cols} campos</span>", unsafe_allow_html=True)

            df_exib = gdf[cols_atrib].copy()
            df_exib = df_exib.astype(str)

            if busca:
                mask = df_exib.apply(lambda col: col.str.contains(busca, case=False, na=False)).any(axis=1)
                df_exib = df_exib[mask]
                st.caption(f"🔍 {len(df_exib)} resultado(s) encontrado(s) para **\"{busca}\"**")

            st.dataframe(df_exib, use_container_width=True, height=450)

        # ── Tab Mapa ──
        with tab2:
            col_op1, col_op2 = st.columns([2, 2])
            with col_op1:
                opcoes_cor = ['(Cor única)'] + cols_atrib
                campo_cor = st.selectbox("🎨 Colorir feições por campo:", opcoes_cor)
                campo_cor = None if campo_cor == '(Cor única)' else campo_cor
            with col_op2:
                if campo_cor:
                    n_vals = gdf[campo_cor].nunique()
                    st.markdown(f"<br><span style='color:#666;font-size:13px'>{n_vals} valores únicos no campo <b>{campo_cor}</b></span>", unsafe_allow_html=True)

            with st.spinner("Gerando mapa..."):
                mapa = gerar_mapa(gdf, campo_cor)

            st.components.v1.html(mapa._repr_html_(), height=520)

        # ── Tab Downloads ──
        with tab3:
            st.markdown("### 📥 Exportar dados")
            data_str = datetime.now().strftime('%Y%m%d_%H%M')

            col_d1, col_d2, col_d3 = st.columns(3)

            # Excel
            with col_d1:
                xlsx_buf = gerar_excel(gdf, arq.name, camada)
                st.download_button(
                    "📊 Baixar tabela em Excel",
                    data=xlsx_buf,
                    file_name=f"{camada}_{data_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.caption("Tabela completa de atributos + aba de metadados")

            # CSV
            with col_d2:
                csv_buf = gdf[cols_atrib].to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
                st.download_button(
                    "📄 Baixar tabela em CSV",
                    data=csv_buf,
                    file_name=f"{camada}_{data_str}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
                st.caption("Compatível com Excel, Power BI e QGIS")

            # Mapa HTML
            with col_d3:
                buf_mapa = io.BytesIO()
                mapa.save(buf_mapa, close_file=False)
                st.download_button(
                    "🗺️ Baixar mapa HTML",
                    data=buf_mapa.getvalue(),
                    file_name=f"mapa_{camada}_{data_str}.html",
                    mime="text/html",
                    use_container_width=True,
                )
                st.caption("Mapa interativo — abre em qualquer navegador")

    except Exception as e:
        st.error(f"Erro ao carregar arquivo: {e}")
        st.exception(e)

else:
    # Tela inicial vazia
    st.markdown("""
    <div style="text-align:center;padding:3rem;color:#888;">
        <div style="font-size:4rem">🗺️</div>
        <div style="font-size:1.2rem;font-weight:bold;color:#1F4E79;margin-top:1rem">
            Faça upload de um arquivo para começar
        </div>
        <div style="margin-top:0.5rem">
            Formatos aceitos: <b>.gpkg</b> · <b>.zip</b> (shapefile ou gpkg)
        </div>
    </div>
    """, unsafe_allow_html=True)

# ── Rodapé ──
st.divider()
st.caption("Águas do Rio | Planejamento e Controle · Visualizador de Camadas Vetoriais")
